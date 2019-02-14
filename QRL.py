from qiskit import QuantumCircuit, ClassicalRegister, QuantumRegister
from qiskit import available_backends, execute, register, get_backend
from qiskit.tools.visualization import circuit_drawer
from qiskit.tools.qi.qi import state_fidelity
from qiskit import Aer
from qiskit import IBMQ
IBMQ.enable_account('<your_api_key>', 'https://quantumexperience.ng.bluemix.net/api')
IBMQ.load_accounts()
IBMQ.backends()

import math
import random 
from math import pi
import openpyxl                                                                        
wb = openpyxl.load_workbook("C:\\Users\\HP\\QRL\\Simulations.xlsx")
sheet = wb.create_sheet()
sheet.title = "Experment Set 1"

for nexp in range(0,1):
    backend = IBMQ.get_backend('ibmq_qasm_simulator')
    shots = 1
    max_credits = 3
   
    sheet.cell(row=1, column=1+4*nexp).value = "Iteration"
    sheet.cell(row=1, column=2+4*nexp).value = "Delta"
    sheet.cell(row=1, column=3+4*nexp).value = "Measurement"
    sheet.cell(row=1, column=4+4*nexp).value = "Fidelity"
    
    delta = pi
    eta = 0.75
    
    x = random.uniform(0, pi)
    y = random.uniform(0, 2*pi)
    z = random.uniform(0, 2*pi)
    
    environment = [math.cos(x/2), complex(math.cos(y), math.sin(y))*math.sin(x/2)]
    count = 0
    a = []
    b = []
    
    q = QuantumRegister(3, 'q')
    c = ClassicalRegister(1)
    qc = QuantumCircuit(q,c)
        
    while count < 50:
        qc.reset(q[0])
        qc.reset(q[1])
        qc.reset(q[2])
        
        qc.u3(x,y,z,q[2])
        for j in range(0,len(a)):
            qc.rx(a[len(a)-1-j],q[0])
            qc.rz(b[len(a)-1-j],q[0])
            qc.rz(-b[j],q[2])
            qc.rx(-a[j],q[2])
        qc.cx(q[2],q[1])
        qc.measure(q[1],c[0])
        
        print(delta)
        backend = IBMQ.get_backend('ibmq_qasm_simulator')
        job = execute(qc, backend=backend, shots=shots, max_credits=max_credits)
        result = job.result()
        result = result.get_counts(qc)
        m = str(result)
        print(m[4])
        sheet.cell(row=count+2, column=1+4*nexp).value = count+1
        sheet.cell(row=count+2, column=2+4*nexp).value = delta
        sheet.cell(row=count+2, column=3+4*nexp).value = m[4]
        
        if(m[4]=='1'):
            a.append(random.uniform(-delta/2, delta/2))
            b.append(random.uniform(-delta/2, delta/2))
            delta = delta/eta
        else:
            delta = delta*eta
        
        q1 = QuantumRegister(1, 'q1')
        qc1 = QuantumCircuit(q1)
        
        for j in range(0,len(a)):
            qc1.rx(a[len(a)-1-j],q1[0])
            qc1.rz(b[len(a)-1-j],q1[0])
    
        backend = Aer.get_backend('statevector_simulator')
        job = execute(qc1, backend)
        qc_state = job.result().get_statevector(qc1)
    
        f = state_fidelity(environment,qc_state)
        print(f)
        sheet.cell(row=count+2, column=4+4*nexp).value = f
        
        wb.save("C:\\Users\\HP\\QRL\\Simulations.xlsx")
        count = count + 1
    
    sheet.cell(row=count+3, column=1+4*nexp).value = "Delta"
    sheet.cell(row=count+3, column=2+4*nexp).value = pi
    sheet.cell(row=count+4, column=1+4*nexp).value = "Eta"
    sheet.cell(row=count+4, column=2+4*nexp).value = 0.75
    sheet.cell(row=count+5, column=1+4*nexp).value = "Theta"
    sheet.cell(row=count+5, column=2+4*nexp).value = x
    sheet.cell(row=count+6, column=1+4*nexp).value = "Phi"
    sheet.cell(row=count+6, column=2+4*nexp).value = y
    wb.save("C:\\Users\\HP\\QRL\\Simulations.xlsx")
    
#QASM_source = qc.qasm()
#print(QASM_source)
#circuit_drawer(qc)
